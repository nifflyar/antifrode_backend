from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from io import BytesIO

class ExcelReportGenerator:
    """Helper class to generate formatted Excel reports."""

    def __init__(self, title: str):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self._set_header_style()

    def _set_header_style(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        self.align_center = Alignment(horizontal="center", vertical="center")

    def write_headers(self, headers: list[str]):
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.align_center

    def write_rows(self, data: list[list]):
        from openpyxl.utils import get_column_letter
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=value)
                
                # Format dates
                if isinstance(value, datetime):
                    cell.value = value.replace(tzinfo=None)
                    cell.number_format = "DD.MM.YYYY HH:MM"

        # Auto-adjust column width
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column].width = adjusted_width

    def get_file_bytes(self) -> bytes:
        output = BytesIO()
        self.wb.save(output)
        return output.getvalue()
