"""Excel-парсер для загрузки транзакций жд-перевозок.

Читает .xlsx файл, валидирует заголовки колонок,
нормализует типы данных и возвращает список RawTransaction.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import openpyxl
import re
from openpyxl.worksheet.worksheet import Worksheet


#  Обязательные колонки (регистро-независимо) 
REQUIRED_COLUMNS: frozenset[str] = frozenset(
    {
        "op_type",
        "op_datetime",
        "amount",
    }
)

#  Маппинг: возможные названия колонок → канонические имена 
COLUMN_ALIASES: dict[str, str] = {
    # op_type
    "тип": "op_type",
    "тип операции": "op_type",
    "op_type": "op_type",
    "type": "op_type",
    "операция": "op_type",
    # op_datetime
    "дата": "op_datetime",
    "дата операции": "op_datetime",
    "op_datetime": "op_datetime",
    "datetime": "op_datetime",
    "дата транзакции": "op_datetime",
    # dep_datetime
    "дата отправления": "dep_datetime",
    "дата/время отправки": "dep_datetime",
    "дата/время отправления": "dep_datetime",
    "dep_datetime": "dep_datetime",
    "departure": "dep_datetime",
    # train
    "поезд": "train_no",
    "train_no": "train_no",
    "номер поезда": "train_no",
    # channel
    "канал": "channel",
    "channel": "channel",
    "канал продажи": "channel",
    "канал продаж": "channel",
    # aggregator
    "агрегатор": "aggregator",
    "aggregator": "aggregator",
    # terminal
    "терминал": "terminal",
    "terminal": "terminal",
    # cashdesk
    "касса": "cashdesk",
    "cashdesk": "cashdesk",
    "кассовый узел": "cashdesk",
    # point_of_sale
    "точка продажи": "point_of_sale",
    "пункт продажи": "point_of_sale",
    "pos": "point_of_sale",
    "point_of_sale": "point_of_sale",
    # amount
    "сумма": "amount",
    "цена": "amount",
    "amount": "amount",
    "стоимость": "amount",
    # fee
    "комиссия": "fee",
    "fee": "fee",
    "сбор": "fee",
    "разные сборы": "fee",
    # fio
    "фио": "fio",
    "fio": "fio",
    "пассажир": "fio",
    "ф.и.о.": "fio",
    "ф.и.о. пассажира": "fio",
    "данные пассажира": "fio",
    # iin
    "иин": "iin",
    "iin": "iin",
    "инн": "iin",
    "номер телефона, иин": "iin",
    "номер тел, иин": "iin",
    # doc_no
    "документ": "doc_no",
    "doc_no": "doc_no",
    "номер документа": "doc_no",
    "серия и номер": "doc_no",
    " документа": "doc_no",
    "номер билета": "doc_no",
    # order_no
    "номер заказа": "order_no",
    "order_no": "order_no",
    # stations
    "станция отправления": "dep_station",
    "станция отправ": "dep_station",
    "станция назначения": "arr_station",
    "станция назна": "arr_station",
    "маршрут": "route",
    # cashdesk
    "касса": "cashdesk",
    "cashdesk": "cashdesk",
    "кассовый узел": "cashdesk",
    "пункт продажи": "cashdesk",
    "pos": "cashdesk",
    # phone
    "телефон": "phone",
    "номер телефона": "phone",
    "контакт": "phone",
    "phone": "phone",
    # source
    "источник": "source",
    "source": "source",
    # tariff_type
    "тип тарифа": "tariff_type",
    "tariff_type": "tariff_type",
    "тип тарифа (льгота)": "tariff_type",
    "льгота": "tariff_type",
    # service_class
    "класс обслуживания": "service_class",
    "service_class": "service_class",
    "класс": "service_class",
    # gender
    "пол": "gender",
    "gender": "gender",
    # branch
    "филиал": "branch",
    "branch": "branch",
    # carrier
    "перевозчик": "carrier",
    "carrier": "carrier",
    # settlement_type
    "тип расчёта": "settlement_type",
    "тип расчета": "settlement_type",
    "settlement_type": "settlement_type",
}


@dataclass
class RawTransaction:
    """Сырая строка из Excel — после парсинга, до domain-конвертации."""

    op_type: str                         # "sale" / "refund"
    op_datetime: datetime
    amount: float
    fee: float = 0.0
    dep_datetime: datetime | None = None
    train_no: str | None = None
    channel: str | None = None
    aggregator: str | None = None
    terminal: str | None = None
    cashdesk: str | None = None
    point_of_sale: str | None = None
    fio: str | None = None
    iin: str | None = None
    doc_no: str | None = None
    order_no: str | None = None
    dep_station: str | None = None
    arr_station: str | None = None
    route: str | None = None
    phone: str | None = None
    tariff_type: str | None = None
    service_class: str | None = None
    gender: str | None = None
    branch: str | None = None
    carrier: str | None = None
    settlement_type: str | None = None
    source: str = "excel_upload"
    # Порядковый номер строки для сообщений об ошибках
    _row_num: int = field(default=0, compare=False, repr=False)


class ExcelParseError(Exception):
    """Ошибка парсинга Excel-файла."""


class ExcelParser:
    """Парсер Excel-файлов с транзакциями.

    Usage::

        parser = ExcelParser()
        transactions = parser.load(file_bytes)
    """

    def __init__(self, sheet_name: str | int = 0) -> None:
        self._sheet_name = sheet_name

    #  Public API 

    def load(self, file_bytes: bytes) -> tuple[list[RawTransaction], list[str]]:
        """Читает Excel из байт и возвращает (список RawTransaction, список ошибок).

        Args:
            file_bytes: Содержимое .xlsx файла.

        Returns:
            Кортеж из (список транзакций, список строк с ошибками).
        """
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelParseError(f"Не удалось открыть файл: {exc}") from exc

        ws: Worksheet = self._get_sheet(wb)
        
        # 1. Сначала ищем заголовки, просматривая первые 20 строк
        # iter_rows в режиме read_only=True работает как генератор
        rows_for_headers = list(ws.iter_rows(max_row=20, values_only=True))
        if not rows_for_headers:
            raise ExcelParseError("Файл пустой — нет строк.")
            
        header_idx, col_map = self._find_headers(rows_for_headers)

        transactions: list[RawTransaction] = []
        errors: list[str] = []

        # 2. Теперь читаем данные начиная со строки ПОСЛЕ заголовков
        # Важно: мы заново открываем итератор, чтобы не хранить все строки в памяти
        # (в read_only режиме iter_rows можно вызывать многократно)
        for row_num, row in enumerate(ws.iter_rows(min_row=header_idx + 2, values_only=True), start=header_idx + 2):
            if all(cell is None for cell in row):
                continue  # пропустить пустые строки
            try:
                tx = self._parse_row(row, col_map, row_num)
                transactions.append(tx)
            except (ValueError, TypeError) as exc:
                errors.append(f"Строка {row_num}: {exc}")

        return transactions, errors

    def validate_columns(self, file_bytes: bytes) -> bool:
        """Быстрая проверка наличия обязательных колонок (без парсинга данных)."""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            ws = self._get_sheet(wb)
            # Берем первые 20 строк для поиска
            rows = list(ws.iter_rows(max_row=20, values_only=True))
            if not rows:
                return False
            self._find_headers(rows)
            return True
        except ExcelParseError:
            return False

    #  Internals 

    def _get_sheet(self, wb: openpyxl.Workbook) -> Worksheet:
        try:
            if isinstance(self._sheet_name, int):
                return wb.worksheets[self._sheet_name]
            return wb[self._sheet_name]
        except (IndexError, KeyError) as exc:
            raise ExcelParseError(
                f"Лист '{self._sheet_name}' не найден. Доступные: {wb.sheetnames}"
            ) from exc

    @classmethod
    def _find_headers(cls, rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
        """Ищет строку заголовков среди первых 20 строк."""
        for row_idx, row in enumerate(rows[:20]):
            headers = [str(h).strip().lower() if h is not None else "" for h in row]
            col_map = cls._build_column_map(headers)
            
            # Если все обязательные колонки найдены, эту строку считаем заголовком
            if not (REQUIRED_COLUMNS - col_map.keys()):
                return row_idx, col_map

        # Если не нашли, проверяем первую строку, чтобы выбросить классическую ошибку
        if rows:
            headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
            col_map = cls._build_column_map(headers)
            cls._validate_required(col_map)

        raise ExcelParseError("Файл пустой — нет данных в строках заголовка.")

    @staticmethod
    def _build_column_map(headers: list[str]) -> dict[str, int]:
        """Строит словарь {canonical_name: col_index} из заголовков."""
        col_map: dict[str, int] = {}
        for idx, header in enumerate(headers):
            canonical = COLUMN_ALIASES.get(header, header)
            if canonical not in col_map:
                col_map[canonical] = idx
        return col_map

    @staticmethod
    def _validate_required(col_map: dict[str, int]) -> None:
        missing = REQUIRED_COLUMNS - col_map.keys()
        if missing:
            raise ExcelParseError(
                f"Отсутствуют обязательные колонки: {', '.join(sorted(missing))}. "
                f"Найдены: {', '.join(sorted(col_map.keys()))}"
            )

    @classmethod
    def _parse_row(
        cls, row: tuple[Any, ...], col_map: dict[str, int], row_num: int
    ) -> RawTransaction:
        def get(name: str) -> Any:
            idx = col_map.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        op_type = cls._parse_op_type(get("op_type"), row_num)
        op_datetime = cls._parse_datetime(get("op_datetime"), "op_datetime", row_num)
        dep_datetime = cls._parse_datetime(get("dep_datetime"), "dep_datetime", row_num, required=False)
        # Для нефинансовых операций (как гашение) сумма может отсутствовать, ставим 0.0
        amount = cls._parse_float(get("amount"), "amount", row_num, default=0.0)
        fee = cls._parse_float(get("fee"), "fee", row_num, default=0.0)

        fio = cls._str_or_none(get("fio"))
        iin = cls._str_or_none(get("iin"))
        
        # Если ИИН в отдельной колонке пуст, попробуем вытащить его из ФИО 
        if not iin and fio:
            iin = cls._extract_iin(fio)

        dep_station = cls._str_or_none(get("dep_station"))
        arr_station = cls._str_or_none(get("arr_station"))
        route = cls._str_or_none(get("route"))
        
        # Генерация маршрута из станций, если самого поля нет (Notebook 6 logic)
        if not route and dep_station and arr_station:
            route = f"{dep_station} -> {arr_station}"

        return RawTransaction(
            op_type=op_type,
            op_datetime=op_datetime,
            dep_datetime=dep_datetime,
            amount=amount,
            fee=fee,
            train_no=cls._str_or_none(get("train_no")),
            channel=cls._str_or_none(get("channel")),
            aggregator=cls._str_or_none(get("aggregator")),
            terminal=cls._str_or_none(get("terminal")),
            cashdesk=cls._str_or_none(get("cashdesk")),
            point_of_sale=cls._str_or_none(get("point_of_sale")),
            fio=fio,
            iin=iin,
            doc_no=cls._str_or_none(get("doc_no")),
            order_no=cls._str_or_none(get("order_no")),
            dep_station=dep_station,
            arr_station=arr_station,
            route=route,
            phone=cls._str_or_none(get("phone")),
            tariff_type=cls._str_or_none(get("tariff_type")),
            service_class=cls._str_or_none(get("service_class")),
            gender=cls._str_or_none(get("gender")),
            branch=cls._str_or_none(get("branch")),
            carrier=cls._str_or_none(get("carrier")),
            settlement_type=cls._str_or_none(get("settlement_type")),
            source=cls._str_or_none(get("source")) or "excel_upload",
            _row_num=row_num,
        )

    #  Type coercions 

    @staticmethod
    def _parse_op_type(value: Any, row_num: int) -> str:
        if value is None:
            raise ValueError(f"op_type пуст в строке {row_num}")
        s = str(value).strip().lower()
        mapping = {
            "sale": "sale", "продажа": "sale", "покупка": "sale",
            "оформление": "sale", "продажа-прин": "sale",
            "подтверждение оплаты": "sale", "оплата": "sale",
            "refund": "refund", "возврат": "refund",
            "redeem": "redeem", "гашение": "redeem",
        }
        result = mapping.get(s)
        if result is None:
            # Вместо ошибки теперь используем 'other', чтобы загружать ВСЕ строки
            return "other"
        return result

    @staticmethod
    def _parse_datetime(
        value: Any, field_name: str, row_num: int, required: bool = True
    ) -> datetime | None:
        if value is None:
            if required:
                raise ValueError(f"{field_name} пуст в строке {row_num}")
            return None
        if isinstance(value, datetime):
            return value
        # Попытка парсинга строк в разных форматах
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
        raise ValueError(
            f"{field_name}='{value}' не удалось преобразовать в дату в строке {row_num}"
        )

    @staticmethod
    def _parse_float(
        value: Any, field_name: str, row_num: int, default: float | None = None
    ) -> float:
        if value is None:
            if default is not None:
                return default
            raise ValueError(f"{field_name} пуст в строке {row_num}")
        try:
            # Чистим строку: убираем пробелы (обычные и неразрывные \u00a0), запятые -> точки
            s = str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".")
            return float(s)
        except (ValueError, TypeError):
            raise ValueError(
                f"{field_name}='{value}' не является числом в строке {row_num}"
            )

    @staticmethod
    def _extract_iin(text: str) -> str | None:
        """Ищет 12 цифр подряд в тексте (ИИН)."""
        # Ищем 12 цифр, возможно после "ИИ" или "##"
        match = re.search(r"(?:ИИ|##|\s)(\d{12})(?:\s|$)", text)
        if not match:
            # Просто любые 12 цифр подряд
            match = re.search(r"(\d{12})", text)
        return match.group(1) if match else None

    @staticmethod
    def _str_or_none(value: Any) -> str | None:
        if value is None:
            return None
        s = str(value).strip()
        return s if s else None
